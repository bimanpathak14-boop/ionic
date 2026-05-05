"""Excel Spreadsheet Handler"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter


class ExcelHandler:
    def __init__(self):
        self.output_dir = os.path.expanduser('~/Documents/PocketAI')
        os.makedirs(self.output_dir, exist_ok=True)

    def create(self, title='Untitled', headers=None, rows=None,
               formulas=None, chart=None, **kwargs):
        """Create a new Excel spreadsheet"""
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel sheet name limit

        # Header styling
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')

        # Add headers
        if headers:
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                ws.column_dimensions[get_column_letter(col)].width = max(len(str(header)) + 4, 12)

        # Add data rows
        if rows:
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-size columns based on ALL data
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = max(adjusted_width, 12)

        # Add formulas
        if formulas:
            for formula in formulas:
                cell = formula.get('cell', 'A1')
                expr = formula.get('formula', '')
                ws[cell] = expr

        # Add chart if requested
        if chart and headers and rows:
            chart_obj = BarChart()
            chart_obj.title = chart.get('title', title)
            chart_obj.style = 10
            data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(rows) + 1, max_col=len(headers))
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(rows) + 1)
            chart_obj.add_data(data_ref, titles_from_data=True)
            chart_obj.set_categories(cats)
            ws.add_chart(chart_obj, 'A' + str(len(rows) + 4))

        # Auto-filter
        if headers:
            ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{len(rows or []) + 1}'

        # Save with unique filename if busy
        base_filename = self._safe_filename(title)
        filename = base_filename + '.xlsx'
        filepath = os.path.join(self.output_dir, filename)
        
        counter = 1
        while os.path.exists(filepath):
            try:
                wb.save(filepath)
                break
            except PermissionError:
                filename = f"{base_filename}_{counter}.xlsx"
                filepath = os.path.join(self.output_dir, filename)
                counter += 1
            except Exception:
                import datetime
                filename = f"{base_filename}_{datetime.datetime.now().strftime('%H%M%S')}.xlsx"
                filepath = os.path.join(self.output_dir, filename)
                wb.save(filepath)
                break
        else:
            wb.save(filepath)
        
        # Open the file automatically
        try:
            os.startfile(filepath)
        except:
            pass

        return {
            'data': {'message': f'Spreadsheet "{title}" created', 'path': filepath},
            'filesCreated': [{'name': filename, 'path': filepath, 'type': 'spreadsheet',
                              'mimeType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                              'size': os.path.getsize(filepath)}],
        }

    def _safe_filename(self, name):
        return "".join(c for c in name if c.isalnum() or c in (' ', '-', '_')).strip()
